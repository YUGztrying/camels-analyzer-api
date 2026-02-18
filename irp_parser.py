"""
IRP Report Excel Parser
-----------------------
Deterministic parser for the structured Excel exported by the IRP Spreading App.
No LLM required — maps exact IFC field names to BankDB columns.

Sheet: "IRP Report"
  Row 1:  "IRP REPORT - MICROFINANCE FORMAT" or "IRP REPORT - BANQUE FORMAT"
  Row 2:  "Company: Acme MFI"
  Row 3:  "Generated: ..."
  Row 5:  "REPORTING PERIODS:"
  Row 6:  "" | "Jan 2023" | "Dec 2023" | "Dec 2024"
  ...
  Sections delimited by "=== ASSETS ===", "=== LIABILITIES & EQUITY ===",
  "=== INCOME STATEMENT ==="

Column layout:
  Col A (idx 0): IFC row label
  Col B (idx 1): period 1 value
  Col C (idx 2): period 2 value
  ...

Values are integers (Math.round); None/missing = null.
"""

import logging
import calendar
from datetime import date, datetime

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Field maps: IFC label → BankDB column
# "last wins": later rows in the sheet override earlier ones for the same
# BankDB field, so totals (which appear after their granular lines) win.
# ---------------------------------------------------------------------------

_ASSETS_MFI = {
    # Cash
    "Cash & Due from Banks (Non Interest Earning)": "cash_reserves_requirements",
    "Cash & Accounts Receivable":                   None,   # sub-total, skip
    # Securities
    "Government Securities (Trading & AFS)":        "investment_securities",
    "Government Securities (Held to Maturity)":     "investment_securities",
    "Other Trading Securities":                     "investment_securities",
    "Total Securities":                             "investment_securities",   # wins
    # Loans
    "Gross Loans":                                  "gross_loans",
    "Less: Allowance for Loan Loss Reserve":        "loan_loss_provisions",    # stored +, negated in post-proc
    # Earning deposits at other banks
    "Interest Bearing Deposits":                    "due_from_banks",
    "Other Earning Assets":                         "due_from_banks",          # wins if present
    # Fixed & other
    "Net Fixed Assets":                             "fixed_assets",
    "Investments in Subsidiaries/Affiliates":       "investment_in_subs_affiliates",
    "Other Non Earning Assets":                     "other_assets",
    "Total Other Assets":                           "other_assets",            # wins
    # Grand total
    "Total Assets":                                 "total_assets",
}

_ASSETS_BANK_EXTRA = {
    # Banks may use a slightly different cash label
    "Cash & Due from Banks":                        "cash_reserves_requirements",
    # Bank loan granulars (all overwritten by Gross Loans total — that label is same)
    "Loans - Corporate":                            "gross_loans",
    "Loans - Retail":                               "gross_loans",
    "Loans - Banks":                                "gross_loans",
    "Loans - Government":                           "gross_loans",
    "Loans - Other":                                "gross_loans",
    "Gross Loans to Customers":                     "gross_loans",             # wins
    # Derivatives / repos → due_from_banks (short-term liquid)
    "Repurchase Agreements (Assets)":               "due_from_banks",
}

_LIABILITIES_MFI = {
    # Deposits
    "Savings Deposits - Interest Bearing":          "deposits",
    "Time Deposits - Interest Bearing":             "deposits",
    "Other Deposits - Non Interest Bearing":        "deposits",
    "Total Deposits":                               "deposits",                # wins
    # Short-term borrowings
    "Short Term Loans Payable - Bank":              "short_term_borrowings",
    "Short Term Loans Payable - Other":             "short_term_borrowings",
    "Current Portion - Long Term Debt":             "short_term_borrowings",
    "Short Term Non-Commercial Borrowing (MFI only)": "short_term_borrowings",
    "Total Short Term Loans":                       "short_term_borrowings",   # wins
    # Other current liabilities
    "Accounts Payable":                             "other_liabilities",
    "Accrued Liabilities":                          "other_liabilities",
    "Total Other Current Liabilities":              "other_liabilities",       # wins
    # Long-term debt
    "Long Term Debt - Bank":                        "long_term_debt",
    "Long Term Debt - IFC":                         "long_term_debt",
    "Other Long Term Debt":                         "long_term_debt",
    "Subordinated Debt":                            "long_term_debt",
    "Long Term Non-Commercial Borrowing (MFI Only)": "long_term_debt",
    "Total Non-Current Liabilities":                "long_term_debt",          # wins
    # Total liabilities
    "Total Liabilities":                            "total_liabilities",
    # Equity components — use "_sum" sentinel for paid_in_capital
    "Common Shares":                                "_paid_in_capital_a",      # special: will be summed
    "Paid-in Capital":                              "_paid_in_capital_b",      # special: will be summed
    "Retained Earnings":                            "retained_earnings",
    "Reserves":                                     "reserves",
    "Donated Equity (MFI only)":                    "reserves",                # treat as reserves
    # Equity totals
    "Total Common Equity":                          "total_equity",
    "TOTAL EQUITY":                                 "total_equity",            # wins
}

_LIABILITIES_BANK_EXTRA = {
    "Deposits from Banks":                          "interbank_liabilities",
    "Repurchase Agreements (Liabilities)":          "interbank_liabilities",
    "Subordinated Debt (Tier II)":                  "long_term_debt",
    "Hybrid Capital":                               "long_term_debt",
    "Preferred Shares":                             "_paid_in_capital_b",      # add to paid-in
}

_INCOME_MFI = {
    # Interest income
    "Loan Interest Income":                         "interest_income",
    "Other Interest Income":                        "interest_income",
    "Total Interest Income":                        "interest_income",         # wins
    # Interest expense (stored as positive)
    "Interest on Deposits":                         "interest_expenses",
    "Interest Expense":                             "interest_expenses",
    "Total Interest Expenses":                      "interest_expenses",       # wins
    # NII — IFC calls it "Net Interest Margin"
    "Net Interest Margin":                          "net_interest_income",
    # Fee income
    "Other Fees - Net":                             "fees_commissions",
    "Total Fee Income":                             "fees_commissions",        # wins
    # Investment income
    "Investment Income (Other than Interest)":      "net_income_investment",
    "Total Investment Income":                      "net_income_investment",   # wins
    # Other income
    "Other Income":                                 "other_revenues",
    "Total Other Income":                           "other_revenues",          # wins
    # Total operating income (= PNB)
    "Total Operating Income":                       "operating_income",
    # Operating expenses
    "Personnel Expenses":                           "wages_salaries",
    "Sales, General & Administrative Expenses":     "other_opex",
    "Depreciation & Amortization Expense":          "fixed_asset_depreciation",
    "All Other Expenses":                           None,                       # skip (part of total)
    "Total Operating Expenses":                     "operating_expenses",      # wins
    # Profit lines
    "Operating Profit before Provision Expenses":   "operating_profit",
    "Net Provision Expenses on Loan Portfolio":     "provision_expenses",      # stored +, abs in post-proc
    "Net Profit From Operations":                   None,                      # derived, skip
    # Non-operating
    "Non Recurring Income":                         "_non_recurring_income",   # special: sign-aware sum
    "Non Recurring Expense":                        "_non_recurring_expense",  # special: sign-aware sum
    "Other Non-Operating Income & (Expenses)":      "non_operating_profit_loss",
    # Tax & net
    "Profit Before Taxes":                          None,                      # skip
    "Income Taxes":                                 "income_tax",
    "Net Income":                                   "net_income",
    # Ignore attribution lines
    "Net Income Attributable to Non Controlling Interests": None,
    "Net Income Attributable to Parent Shareholders": None,
    "Total Comprehensive Income":                   None,
    "TCI Attributable to Parent Shareholders":      None,
}

_INCOME_BANK_EXTRA = {
    # Bank-specific income items
    "Net Gains/Losses from Trading (Securities/FX/Derivatives)": "fvtpl_changes",
    "Net Fees & Commissions":                       "fees_commissions",        # wins over "Other Fees - Net"
    "Non-Interest Income":                          "non_interest_income_commissions",
    "Pre-tax Profit":                               None,                      # bank calls it this; skip
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def is_irp_report(file_path: str) -> bool:
    """Return True if this file is an IRP Report Excel (not a raw bank document)."""
    if not file_path.lower().endswith((".xlsx", ".xls")):
        return False
    try:
        from openpyxl import load_workbook
        # No data_only: strings starting with '=' must not be misread as formulas
        wb = load_workbook(file_path, read_only=True)
        if "IRP Report" not in wb.sheetnames:
            return False
        ws = wb["IRP Report"]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        return bool(first_row and first_row[0] and "IRP REPORT" in str(first_row[0]).upper())
    except Exception:
        return False


def parse_irp_excel(file_path: str) -> list[dict]:
    """
    Parse an IRP Report Excel file.

    Returns a list of dicts sorted oldest→newest (one dict per period column).
    Each dict has all financial fields ready to populate a BankDB row, plus:
      - "name": company name
      - "fiscal_year": e.g. "2023"
      - "period_label": e.g. "Dec 2023"
      - "period_end_date": date object (last day of that month)
      - "institution_type": "bank" | "mfi"
      - "currency": "XOF"
    """
    from openpyxl import load_workbook
    # No data_only: section headers like "=== ASSETS ===" start with '='
    # and must not be misread as cached-formula cells returning None.
    # Data values in IRP files are always plain integers, never formulas.
    wb = load_workbook(file_path)
    ws = wb["IRP Report"]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("IRP Report sheet is empty")

    # ── Detect institution type ────────────────────────────────────────────
    title = str(rows[0][0] or "").upper()
    is_bank = "BANK" in title or "BANQUE" in title
    institution_type = "bank" if is_bank else "mfi"
    logger.info("IRP parser: institution_type=%s", institution_type)

    # ── Extract company name ───────────────────────────────────────────────
    company_name = "Unknown"
    for row in rows[:6]:
        cell = str(row[0] or "").strip()
        if cell.lower().startswith("company:"):
            company_name = cell.split(":", 1)[1].strip()
            break
    logger.info("IRP parser: company=%s", company_name)

    # ── Find period headers ────────────────────────────────────────────────
    # The period row has blanks/text in col A and date strings ("Jan 2023") from col B onward.
    period_labels: list[str] = []
    for row in rows[:12]:
        cells = [str(c or "").strip() for c in row]
        candidates = [c for c in cells[1:] if c and _parse_period_date(c) is not None]
        if candidates:
            period_labels = candidates
            break

    if not period_labels:
        raise ValueError("No period headers found in IRP Report (expected 'Jan 2023' style)")

    num_periods = len(period_labels)
    period_dates = [_parse_period_date(lbl) for lbl in period_labels]
    fiscal_years = [str(d.year) if d else "N/A" for d in period_dates]
    logger.info("IRP parser: periods=%s", period_labels)

    # ── Build merged field maps ────────────────────────────────────────────
    assets_map = dict(_ASSETS_MFI)
    liab_map = dict(_LIABILITIES_MFI)
    income_map = dict(_INCOME_MFI)
    if is_bank:
        assets_map.update(_ASSETS_BANK_EXTRA)
        liab_map.update(_LIABILITIES_BANK_EXTRA)
        income_map.update(_INCOME_BANK_EXTRA)

    # ── Scan rows ──────────────────────────────────────────────────────────
    SECTION_ASSETS = "assets"
    SECTION_LIAB   = "liabilities"
    SECTION_INCOME = "income"

    current_section = None
    # One raw accumulator per period
    period_raw: list[dict] = [{} for _ in range(num_periods)]

    _SKIP_LABELS = {"description", ""}
    _SECTION_STARTERS = {
        "=== assets ===": SECTION_ASSETS,
        "=== liabilities & equity ===": SECTION_LIAB,
        "=== income statement ===": SECTION_INCOME,
    }

    for row in rows:
        col_a = str(row[0] or "").strip()
        col_a_lower = col_a.lower()

        # Detect section delimiter
        matched_section = None
        for key, sec in _SECTION_STARTERS.items():
            if key in col_a_lower:
                matched_section = sec
                break
        if matched_section:
            current_section = matched_section
            continue

        # Skip structural / header rows
        if (not col_a
                or col_a_lower in _SKIP_LABELS
                or col_a_lower.startswith("irp report")
                or col_a_lower.startswith("company:")
                or col_a_lower.startswith("generated:")
                or col_a_lower.startswith("reporting")
                or current_section is None):
            continue

        # Choose field map for this section
        if current_section == SECTION_ASSETS:
            field_map = assets_map
        elif current_section == SECTION_LIAB:
            field_map = liab_map
        else:
            field_map = income_map

        if col_a not in field_map:
            continue

        db_field = field_map[col_a]
        if db_field is None:
            continue  # explicitly skipped

        # Read one value per period
        for p_idx in range(num_periods):
            col_idx = p_idx + 1          # col B = index 1, col C = index 2, …
            if col_idx >= len(row):
                continue
            raw = row[col_idx]
            if raw is None:
                continue
            try:
                val = float(raw)
            except (TypeError, ValueError):
                continue

            period_raw[p_idx][db_field] = val

    # ── Post-process each period ───────────────────────────────────────────
    results: list[dict] = []

    for p_idx, raw in enumerate(period_raw):
        data = dict(raw)

        # Combine Common Shares + Paid-in Capital into paid_in_capital
        a = data.pop("_paid_in_capital_a", None)
        b = data.pop("_paid_in_capital_b", None)
        combined = _safe_sum(a, b)
        if combined is not None:
            data["paid_in_capital"] = combined

        # Combine non-recurring items into non_operating_profit_loss if not already set
        nr_inc = data.pop("_non_recurring_income", None)
        nr_exp = data.pop("_non_recurring_expense", None)
        if "non_operating_profit_loss" not in data:
            # non_recurring_expense is a cost → subtract
            nr_exp_neg = (-abs(nr_exp)) if nr_exp is not None else None
            nop = _safe_sum(nr_inc, nr_exp_neg)
            if nop is not None:
                data["non_operating_profit_loss"] = nop

        # loan_loss_provisions: IFC stores as positive (it's a "Less:" line).
        # BankDB convention: negative.
        llp = data.get("loan_loss_provisions")
        if llp is not None and llp > 0:
            data["loan_loss_provisions"] = -llp

        # interest_expenses: must be positive (our convention)
        ie = data.get("interest_expenses")
        if ie is not None and ie < 0:
            data["interest_expenses"] = abs(ie)

        # provision_expenses: must be positive
        pe = data.get("provision_expenses")
        if pe is not None and pe < 0:
            data["provision_expenses"] = abs(pe)

        # income_tax: must be positive
        tax = data.get("income_tax")
        if tax is not None and tax < 0:
            data["income_tax"] = abs(tax)

        # Mirror net_income → net_profit (BankDB has both; net_profit is in equity section)
        if "net_income" in data:
            data.setdefault("net_profit", data["net_income"])

        entry = {
            "name":             company_name,
            "fiscal_year":      fiscal_years[p_idx],
            "period_label":     period_labels[p_idx],
            "period_end_date":  period_dates[p_idx],
            "currency":         "XOF",
            "institution_type": institution_type,
            **data,
        }
        results.append(entry)
        logger.info(
            "IRP parser period %s: total_assets=%s  net_income=%s",
            period_labels[p_idx],
            data.get("total_assets"),
            data.get("net_income"),
        )

    # Sort oldest → newest so averaging works correctly (prev = results[i-1])
    results.sort(key=lambda r: r["period_end_date"] or date.min)
    return results


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_period_date(s: str) -> date | None:
    """
    Parse 'Jan 2023', 'Dec 2024' (toLocaleDateString en-US short) into the
    last calendar day of that month.
    """
    try:
        dt = datetime.strptime(s.strip(), "%b %Y")
        last_day = calendar.monthrange(dt.year, dt.month)[1]
        return date(dt.year, dt.month, last_day)
    except ValueError:
        return None


def _safe_sum(*vals):
    """Sum non-None values; return None if all are None."""
    non_none = [v for v in vals if v is not None]
    return sum(non_none) if non_none else None
