import anthropic
import os
import logging
from dotenv import load_dotenv
import json
from PyPDF2 import PdfReader

load_dotenv()
logger = logging.getLogger(__name__)

_api_key = os.getenv("ANTHROPIC_API_KEY")
if not _api_key:
    logger.warning("ANTHROPIC_API_KEY not set — document extraction will fail")

client = anthropic.Anthropic(api_key=_api_key)

def extract_bank_data_from_file(file_path: str) -> dict:
    """
    Extract structured financial data from a bank document (PDF or image).
    Supports text-based PDFs, scanned PDFs (OCR), and direct images.
    """

    prompt = """
You are an expert banking financial analyst specializing in the UEMOA (West African) zone.

Your mission: Extract ALL balance sheet and income statement data from this document.

CRITICAL INSTRUCTIONS:
- The document may contain 30+ pages, but only 3-5 pages matter
- Focus on: "BILAN", "BALANCE SHEET", "COMPTE DE RESULTAT", "INCOME STATEMENT"
- Ignore detailed notes/annexes
- Amounts may be in thousands (KCFA, MXOF) or millions

================================================================================
BALANCE SHEET — ASSETS (look for these labels or equivalents):
================================================================================

- cash_reserves_requirements -> "Caisse et Banque Centrale" / "Cash & Central Bank"
- due_from_banks -> "Correspondants bancaires" / "Due from Banks"
- investment_securities -> "Titres de placement" / "Investment Securities"
- gross_loans -> "Credits a la clientele" (GROSS, before provisions) / "Gross Loans"
- loan_loss_provisions -> "Provisions pour creances douteuses" / "Loan Loss Provisions" (NEGATIVE value)
- foreclosed_assets -> "Actifs saisis" / "Foreclosed Assets"
- fixed_assets -> "Immobilisations" / "Fixed Assets"
- other_assets -> "Autres actifs"
- total_assets -> "TOTAL ACTIF" / "TOTAL ASSETS" (REQUIRED)

NPL DATA:
- npls_mn -> "Creances en souffrance" / "NPL" / "Stage 3 receivables" — if missing, use absolute value of loan_loss_provisions
- llr_mn -> "Provisions sur creances" / "Loan Loss Reserves" — if missing, use absolute value of loan_loss_provisions

================================================================================
BALANCE SHEET — LIABILITIES:
================================================================================

- deposits -> "Depots de la clientele" / "Customer Deposits"
- short_term_borrowings -> "Emprunts a court terme" / "Short-Term Borrowings" / "Dettes a court terme"
- long_term_debt -> "Emprunts a long terme" / "Long-Term Debt" / "Dettes a moyen et long terme"
- interbank_liabilities -> "Dettes envers les etablissements de credit"
- other_liabilities -> "Autres passifs"
- total_liabilities -> "TOTAL PASSIF" / "Total Liabilities"

================================================================================
BALANCE SHEET — EQUITY:
================================================================================

- paid_in_capital -> "Capital social" / "Share Capital"
- reserves -> "Reserves"
- retained_earnings -> "Report a nouveau" / "Retained Earnings"
- net_profit -> "Resultat de l'exercice" / "Net Profit"
- total_equity -> "CAPITAUX PROPRES" / "Total Equity" (REQUIRED)

================================================================================
INCOME STATEMENT:
================================================================================

- interest_income -> "Produits d'interets" / "Interest Income"
- interest_expenses -> "Charges d'interets" / "Interest Expenses" (as positive number)
- net_interest_income -> "Marge d'interet" / "Net Interest Income" (= Interest Income MINUS Interest Expenses ONLY)

NON-INTEREST INCOME (extract as much detail as possible):
- fees_commissions -> "Commissions" / "Fees & Commissions"
- net_sales -> "Ventes nettes" / "Net Sales"
- dividends -> "Dividendes recus" / "Dividends"
- fvtpl_changes -> "Variation de juste valeur" / "FVTPL Changes"
- securitization_gains -> "Gains de titrisation" / "Securitization Gains"
- provisions_no_longer_required -> "Reprises de provisions" / "Provisions No Longer Required"
- share_profit_associates -> "Quote-part de resultat des societes mises en equivalence"
- other_revenues -> "Autres produits" / "Other Revenues"
- gain_acquisition_subsidiaries -> "Gain sur acquisition de filiales"
- non_interest_income_commissions -> aggregate of all non-interest income if individual items not available

OPERATING EXPENSES (extract as much detail as possible):
- wages_salaries -> "Charges de personnel" / "Wages & Salaries"
- other_opex -> "Autres charges d'exploitation" / "Other Operating Expenses"
- intangible_amortization -> "Amortissement des immobilisations incorporelles"
- fixed_asset_depreciation -> "Amortissement des immobilisations corporelles" / "Depreciation"
- operating_expenses -> aggregate total operating expenses

- net_income_investment -> "Revenus de placements" / "Investment Income" (income from securities/investments)
- other_net_income -> "Autres produits nets" / "Other Net Income"
- operating_income -> "Produit Net Bancaire" / "PNB" / "Operating Income" / "Total Banking Income" (= NII + commissions + all non-interest income, BEFORE operating expenses)
- operating_profit -> "Resultat d'exploitation" / "Operating Profit"

OTHER P&L:
- provision_expenses -> "Dotations aux provisions" / "Expected Credit Losses (ECL)" / "Provision Expenses"
- provisions_formed -> "Provisions constituees" / "Provisions Formed"
- impairment_financial_assets -> "Depreciation des actifs financiers" / "Impairment"
- fx_exchange -> "Gains/pertes de change" / "Foreign Currency Exchange"
- non_operating_profit_loss -> "Resultat hors exploitation"
- income_tax -> "Impots" / "Income Tax"
- net_income -> "RESULTAT NET" / "NET INCOME" (REQUIRED)

================================================================================
PRUDENTIAL RATIOS (if already computed in the document):
================================================================================

- car_regulatory -> "Ratio de solvabilite" / "CAR" (format: 13.42 for 13.42%)
- car_bank_reported -> "CAR declare par la banque"
- npl_ratio_reported -> "Taux de creances en souffrance" / "NPL Ratio" (format: 5.2 for 5.2%)
- coverage_ratio_reported -> "Taux de couverture" / "Coverage Ratio" (format: 95.5 for 95.5%)
- roe_reported -> "ROE" / "Return on Equity" (format: 28.51 for 28.51%)
- roa_reported -> "ROA" / "Return on Assets" (format: 2.59 for 2.59%)
- cost_income_reported -> "Coefficient d'exploitation" / "Cost to Income" (format: 45.09 for 45.09%)

================================================================================
OUTPUT FORMAT (JSON ONLY):
================================================================================

Return ONLY valid JSON. Use null for missing values (never guess or invent numbers).
All amounts as ABSOLUTE values except loan_loss_provisions (negative).

{
    "name": "Bank Name",
    "country": "Country",
    "fiscal_year": "2023",
    "currency": "XOF",
    "total_assets": 700000,
    "cash_reserves_requirements": 50000,
    "due_from_banks": 20000,
    "investment_securities": 100000,
    "gross_loans": 500000,
    "loan_loss_provisions": -25000,
    "foreclosed_assets": null,
    "fixed_assets": 30000,
    "other_assets": 25000,
    "deposits": 550000,
    "short_term_borrowings": null,
    "long_term_debt": null,
    "interbank_liabilities": null,
    "other_liabilities": 50000,
    "total_liabilities": 600000,
    "paid_in_capital": 60000,
    "reserves": 30000,
    "retained_earnings": 5000,
    "net_profit": 5000,
    "total_equity": 100000,
    "interest_income": 60000,
    "interest_expenses": 20000,
    "net_interest_income": 40000,
    "fees_commissions": 8000,
    "net_sales": null,
    "dividends": null,
    "fvtpl_changes": null,
    "securitization_gains": null,
    "provisions_no_longer_required": null,
    "share_profit_associates": null,
    "other_revenues": 2000,
    "gain_acquisition_subsidiaries": null,
    "non_interest_income_commissions": 10000,
    "net_income_investment": null,
    "other_net_income": null,
    "wages_salaries": 15000,
    "other_opex": 10000,
    "intangible_amortization": 2000,
    "fixed_asset_depreciation": 3000,
    "operating_expenses": 30000,
    "operating_income": 50000,
    "operating_profit": 20000,
    "provision_expenses": 10000,
    "provisions_formed": null,
    "impairment_financial_assets": null,
    "fx_exchange": null,
    "non_operating_profit_loss": null,
    "income_tax": 5000,
    "net_income": 5000,
    "car_regulatory": 14.3,
    "car_bank_reported": null,
    "npls_mn": 25000,
    "llr_mn": 25000,
    "npl_ratio_reported": 4.55,
    "coverage_ratio_reported": 95.5,
    "roe_reported": 28.51,
    "roa_reported": 2.59,
    "cost_income_reported": 45.09
}
"""

    # --- Process by file type ---

    if file_path.lower().endswith(('.xlsx', '.xls')):
        logger.info("Processing Excel file...")
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)
        text = ""
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text += f"\n\n{'='*80}\nSHEET: {sheet_name}\n{'='*80}\n\n"
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = "\t".join(cells).strip()
                if line:
                    text += line + "\n"

        logger.info(f"Excel extracted: {len(text)} characters from {len(wb.sheetnames)} sheet(s)")

        if len(text.strip()) < 50:
            raise Exception("Excel file appears empty or unreadable")

        message = client.messages.create(
            model="claude-3-5-haiku-20241022",
            max_tokens=4096,
            messages=[{
                "role": "user",
                "content": f"{prompt}\n\n{'='*80}\nDOCUMENT (EXCEL SPREADSHEET):\n{'='*80}\n\n{text[:100000]}"
            }]
        )

    elif file_path.lower().endswith('.pdf'):
        logger.info("Processing PDF file...")

        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text()

        if len(text.strip()) < 100:
            # Scanned PDF -> OCR all pages
            logger.info("Scanned PDF detected - running OCR on all pages...")
            from pdf2image import convert_from_path
            import pytesseract

            images = convert_from_path(file_path, dpi=150)
            if not images:
                raise Exception("Failed to convert PDF to images")

            logger.info(f"{len(images)} page(s) converted")

            full_text = ""
            for i, img in enumerate(images):
                logger.info(f"  OCR page {i+1}/{len(images)}...")
                try:
                    page_text = pytesseract.image_to_string(img, lang='fra+eng', config='--psm 6')
                    full_text += f"\n\n{'='*80}\nPAGE {i+1}\n{'='*80}\n\n{page_text}"
                    logger.debug(f"OCR page {i+1} OK ({len(page_text)} chars)")
                except Exception as e:
                    logger.error(f"OCR error on page {i+1}: {e}")

            logger.info(f"OCR complete: {len(full_text)} total characters")

            message = client.messages.create(
                model="claude-3-5-haiku-20241022",
                max_tokens=4096,
                messages=[{
                    "role": "user",
                    "content": f"{prompt}\n\n{'='*80}\nDOCUMENT (OCR EXTRACTED):\n{'='*80}\n\n{full_text[:100000]}"
                }]
            )
        else:
            # Text-based PDF
            logger.info(f"Text PDF ({len(text)} characters)")

            message = client.messages.create(
                model="claude-3-5-haiku-20241022",
                max_tokens=4096,
                messages=[{
                    "role": "user",
                    "content": f"{prompt}\n\n{'='*80}\nDOCUMENT:\n{'='*80}\n\n{text[:100000]}"
                }]
            )
    else:
        # Direct image (JPG/PNG)
        logger.info(f"Image file: {file_path}")
        from PIL import Image
        import pytesseract

        img = Image.open(file_path)
        image_text = pytesseract.image_to_string(img, lang='fra+eng', config='--psm 6')
        logger.info(f"OCR extracted: {len(image_text)} characters")

        message = client.messages.create(
            model="claude-3-5-haiku-20241022",
            max_tokens=4096,
            messages=[{
                "role": "user",
                "content": f"{prompt}\n\n{'='*80}\nDOCUMENT (OCR EXTRACTED):\n{'='*80}\n\n{image_text}"
            }]
        )

    # --- Parse response ---

    response_text = message.content[0].text
    json_str = response_text.strip()

    if "```json" in response_text:
        json_str = response_text.split("```json")[1].split("```")[0].strip()
    elif "```" in response_text:
        json_str = response_text.split("```")[1].split("```")[0].strip()
    else:
        start = response_text.find('{')
        end = response_text.rfind('}') + 1
        if start != -1 and end > start:
            json_str = response_text[start:end]

    try:
        extracted_data = json.loads(json_str)
        logger.info(f"Extracted: {extracted_data.get('name', 'N/A')} - {extracted_data.get('fiscal_year', 'N/A')}")

        # Validate the extracted data
        validation_errors = validate_extracted_data(extracted_data)
        if validation_errors:
            logger.warning(f"Validation warnings: {validation_errors}")
            extracted_data["_validation_warnings"] = validation_errors

        return extracted_data
    except json.JSONDecodeError as e:
        raise Exception(f"Invalid JSON from Claude: {str(e)}")


def validate_extracted_data(data: dict) -> list[str]:
    """
    Validate LLM-extracted financial data for sanity.
    Returns a list of warning/error strings. Empty list = all good.
    """
    errors = []

    # --- Required fields ---
    required_positive = ["total_assets", "total_equity"]
    for field in required_positive:
        val = data.get(field)
        if val is None:
            errors.append(f"Missing required field: {field}")
            continue
        try:
            val = float(val)
        except (ValueError, TypeError):
            errors.append(f"{field} is not a valid number: {val}")
            continue
        if val < 0:
            errors.append(f"{field} should not be negative: {val}")

    # --- Cross-field sanity checks ---
    ta = _to_float(data.get("total_assets"))
    te = _to_float(data.get("total_equity"))
    gl = _to_float(data.get("gross_loans"))
    npls = _to_float(data.get("npls_mn"))
    ni = _to_float(data.get("net_income"))

    if ta and te and te > ta:
        errors.append(f"Total equity ({te}) exceeds total assets ({ta}) — likely an extraction error")

    if gl and npls and npls > gl:
        errors.append(f"NPLs ({npls}) exceed gross loans ({gl}) — likely an extraction error")

    if ta and ni and abs(ni) > ta:
        errors.append(f"Net income ({ni}) exceeds total assets ({ta}) — likely an extraction error")

    # --- Check that interest_expenses is positive (our convention) ---
    ie = _to_float(data.get("interest_expenses"))
    if ie is not None and ie < 0:
        errors.append(f"interest_expenses should be positive (got {ie}); will take absolute value")

    return errors


def _to_float(val) -> float | None:
    """Safely convert to float, return None on failure."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None
